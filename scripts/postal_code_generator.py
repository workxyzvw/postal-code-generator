import pandas as pd
import os
import requests
from bs4 import BeautifulSoup
import time
from tqdm import tqdm
from openpyxl import load_workbook  # Untuk styling batch Excel
from openpyxl.styles import PatternFill  # Untuk styling batch Excel
import re  # Import modul re untuk regular expression
import random  # Import modul random untuk jeda acak

try:
    import cloudscraper  # Import cloudscraper

    CLOUDSCAPER_AVAILABLE = True
except ImportError:
    CLOUDSCAPER_AVAILABLE = False
    print(
        "--- PERINGATAN: Library cloudscraper tidak terinstal. Scraping nomor.net akan menggunakan requests biasa (kemungkinan gagal karena Cloudflare). ---")
    print("--- Untuk mencoba melewati Cloudflare di nomor.net, jalankan: pip install cloudscraper ---")

# Path ke file Excel
try:
    base_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    base_dir = os.getcwd()

input_file = os.path.join(base_dir, "../data/village_postal_code.xlsx")

# --- KONFIGURASI PENGGUNA UNTUK MANUAL CHUNKING ---
# Setiap orang yang menjalankan skrip perlu menyesuaikan tiga variabel ini.
MY_PROCESS_ID = "chief"  # Contoh: "chief", "teman1", "orang_1", "bagian_a", dll.
MY_START_ROW_ABSOLUTE = 620  # Indeks baris Python (0-based) dari file Excel ASLI untuk MULAI proses
MY_END_ROW_EXCLUSIVE_ABSOLUTE = 13857  # Indeks baris Python (0-based) dari file Excel ASLI untuk BERHENTI (eksklusif)
# Contoh: Untuk memproses baris 1-10000 di Excel, set MY_START_ROW_ABSOLUTE = 0, MY_END_ROW_EXCLUSIVE_ABSOLUTE = 10000
# Contoh: Untuk Chief (lanjutan dari 620 data):
# MY_PROCESS_ID = "chief"
# MY_START_ROW_ABSOLUTE = 620
# MY_END_ROW_EXCLUSIVE_ABSOLUTE = 28334
# -------------------------------------------------

BATCH_SIZE = 20
MAX_RETRIES_NOMOR_NET = 3
RETRY_DELAY_NOMOR_NET = 5

# --- Nama File Output & Folder Batch (dengan penanda chunk) ---
output_file_suffix = f"_part_{MY_PROCESS_ID}_{MY_START_ROW_ABSOLUTE}-{MY_END_ROW_EXCLUSIVE_ABSOLUTE - 1}"
output_file = os.path.join(base_dir, f"../data/village_postal_code_enhanced_v14{output_file_suffix}.xlsx")
batches_dir = os.path.join(base_dir, f"../data/batches_enhanced_v14{output_file_suffix}")

# Placeholder strings constants
PH_PROV = "provinsi-kosong"
PH_REG = "kabupaten-kosong"
PH_DIST = "kecamatan-kosong"
PH_VIL = "desa-kosong"
PH_VIL_NOMOR_QUERY_PART = PH_VIL.replace('-', ' ')
PH_DIST_NOMOR_QUERY_PART = PH_DIST.replace('-', ' ')
ENCODED_PH_VIL_NOMOR = requests.utils.quote(PH_VIL_NOMOR_QUERY_PART)
ENCODED_PH_DIST_NOMOR = requests.utils.quote(PH_DIST_NOMOR_QUERY_PART)

# --- DEBUGGING FLAG ---
DEBUG_TARGET_VILLAGE = "gunong pulo"
DEBUG_TARGET_DISTRICT = "kluet utara"
ENABLE_DETAILED_DEBUG = False

print(f"--- DEBUG: Base directory (base_dir): {base_dir} ---")
print(f"--- DEBUG: Target input file: {input_file} ---")
print(f"--- DEBUG (ID: {MY_PROCESS_ID}): Target output file: {output_file} ---")
print(f"--- DEBUG (ID: {MY_PROCESS_ID}): Target batches directory: {batches_dir} ---")

try:
    if not os.path.exists(batches_dir):
        os.makedirs(batches_dir)
        print(f"Folder {batches_dir} telah berhasil dibuat.")
    else:
        print(f"Folder {batches_dir} sudah ada.")
except Exception as e:
    print(f"--- ERROR: Gagal membuat atau mengakses folder batches: {batches_dir} ---")
    print(f"Error: {e}")
    exit(1)

try:
    df_full = pd.read_excel(input_file, sheet_name="villages")
    print(f"File {input_file} berhasil dibaca. Jumlah baris total di Excel: {len(df_full)}")

    # Logika untuk memproses potongan data sesuai konfigurasi manual chunk
    actual_start_row = max(0, MY_START_ROW_ABSOLUTE)
    actual_end_row = min(len(df_full), MY_END_ROW_EXCLUSIVE_ABSOLUTE)

    if actual_start_row < actual_end_row:
        print(
            f"--- INFO (ID: {MY_PROCESS_ID}): Akan memproses baris absolut dari {actual_start_row} hingga {actual_end_row - 1} ---")
        df = df_full.iloc[actual_start_row:actual_end_row].reset_index(drop=True)
        print(f"--- INFO (ID: {MY_PROCESS_ID}): Jumlah baris yang akan diproses oleh instance ini: {len(df)} ---")
    else:
        print(f"--- INFO (ID: {MY_PROCESS_ID}): Rentang baris tidak valid atau tidak ada data untuk diproses. "
              f"Start Absolut: {MY_START_ROW_ABSOLUTE}, End Absolut: {MY_END_ROW_EXCLUSIVE_ABSOLUTE}, Total Data di File: {len(df_full)} ---")
        df = pd.DataFrame()  # Buat DataFrame kosong agar skrip selesai dengan baik

except FileNotFoundError:
    print(f"Error: File {input_file} tidak ditemukan. Pastikan file ada di folder yang benar.")
    exit(1)
except Exception as e:
    print(f"Error saat membaca file Excel atau membagi chunk: {str(e)}")
    print("Pastikan nama sheet (jika ada) sudah benar, misal 'villages'.")
    exit(1)

if df.empty:
    print(
        f"Tidak ada data untuk diproses oleh ID: {MY_PROCESS_ID} (setelah slicing atau file asli kosong). Skrip akan berhenti.")
    exit(0)


# Fungsi buat format nama ke URL
def format_name(text, is_regency=False, placeholder_value="-"):
    if pd.isna(text) or str(text).strip() == "":
        return placeholder_value

    processed_text = str(text).lower().strip()

    def clean_and_hyphenate_name_part(name_part_str, default_placeholder_for_empty):
        if not name_part_str:
            return default_placeholder_for_empty
        name_part_str_no_dots = name_part_str.replace(".", "")
        cleaned = ''.join(char for char in name_part_str_no_dots if char.isalnum() or char == ' ' or char == '-')
        hyphenated = cleaned.strip().replace(" ", "-")
        final_part = re.sub(r'-+', '-', hyphenated)
        final_part = re.sub(r'^-+|-+$', '', final_part)
        return final_part if final_part else default_placeholder_for_empty

    if is_regency:
        if processed_text.startswith("kab."):
            name_part = processed_text[4:].strip()
            hyphenated_name_part = clean_and_hyphenate_name_part(name_part, placeholder_value)
            return f"kabupaten-{hyphenated_name_part}" if hyphenated_name_part != placeholder_value else placeholder_value

        elif processed_text.startswith("kota"):
            if len(processed_text) > 4 and processed_text[4] == ' ':
                name_part = processed_text[5:].strip()
            elif len(processed_text) > 4 and processed_text[4].isalnum():
                name_part = processed_text[4:].strip()
            elif len(processed_text) == 4:
                name_part = ""
            else:
                name_part = processed_text.replace("kota", "", 1).strip()

            hyphenated_name_part = clean_and_hyphenate_name_part(name_part, placeholder_value)
            return f"kota-{hyphenated_name_part}" if hyphenated_name_part != placeholder_value else placeholder_value

    return clean_and_hyphenate_name_part(processed_text, placeholder_value)


# Fungsi buat generate URL kodepos.co.id
def generate_kodepos_url(row):
    prov = format_name(row.get("Provinsi (Province)"), placeholder_value=PH_PROV)
    reg = format_name(row.get("Kabupaten (Regency)"), is_regency=True, placeholder_value=PH_REG)
    dist = format_name(row.get("Kecamatan (District)"), placeholder_value=PH_DIST)
    vil = format_name(row.get("Nama Desa (Village Name)"), placeholder_value=PH_VIL)

    if prov == PH_PROV or reg == PH_REG or dist == PH_DIST or vil == PH_VIL:
        return "URL tidak dapat dibuat (data tidak lengkap)"
    return f"https://kodepos.co.id/kodepos/{prov}/{reg}/{dist}/{vil}"


# Fungsi buat generate URL nomor.net
def generate_nomor_url(row):
    vil_original = row.get("Nama Desa (Village Name)")
    dist_original = row.get("Kecamatan (District)")

    vil = format_name(vil_original, placeholder_value=PH_VIL)
    dist = format_name(dist_original, placeholder_value=PH_DIST)

    if vil == PH_VIL or dist == PH_DIST:
        return "URL tidak dapat dibuat (data tidak lengkap)"

    vil_query_part = vil.replace('-', ' ')
    dist_query_part = dist.replace('-', ' ')

    vil_url = requests.utils.quote(vil_query_part)
    dist_url = requests.utils.quote(dist_query_part)

    generated_url = f"https://www.nomor.net/_kodepos.php?_i=cari-kodepos&jobs={vil_url}%20{dist_url}"
    return generated_url


# Fungsi buat scraping kode pos dari kodepos.co.id
def scrape_kodepos(url):
    if not url or not isinstance(url, str) or not url.startswith("http") or "URL tidak dapat dibuat" in url:
        return None
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        time.sleep(random.uniform(0.5, 1.5))
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        postal_code_tag_td = soup.find("td", itemprop="postalCode")
        if postal_code_tag_td and postal_code_tag_td.string:
            code = postal_code_tag_td.string.strip()
            if code.isdigit() and len(code) == 5: return code

        href_pattern = re.compile(r"/kodepos/(\d{5})$")
        for a_tag in soup.find_all("a", href=True):
            match = href_pattern.search(a_tag['href'])
            if match:
                code_from_href = match.group(1)
                if code_from_href.isdigit() and len(code_from_href) == 5: return code_from_href
            if a_tag.string:
                text_in_a = a_tag.string.strip()
                if text_in_a.isdigit() and len(text_in_a) == 5: return text_in_a

        possible_tags = soup.find_all(['span', 'div', 'strong', 'b', 'td'])
        for tag in possible_tags:
            if tag.string:
                text = tag.string.strip()
                if text.isdigit() and len(text) == 5:
                    parent_text = tag.parent.get_text(separator=" ", strip=True).lower() if tag.parent else ""
                    if "kode pos" in parent_text or "postal code" in parent_text or len(parent_text) < 50:
                        return text
        return None
    except requests.exceptions.HTTPError as http_err:
        if http_err.response.status_code == 404:
            return "Halaman tidak ditemukan (404)"
        return f"Error HTTP ({http_err.response.status_code})"
    except requests.exceptions.RequestException:
        return "Error Request"
    except Exception:
        return "Error Lainnya"

    # Fungsi buat scraping kode pos dari nomor.net (MENGGUNAKAN cloudscraper dengan retry)


def scrape_nomor(url, is_debug_target=False):
    if not url or not isinstance(url, str) or not url.startswith("http") or "URL tidak dapat dibuat" in url:
        if is_debug_target and ENABLE_DETAILED_DEBUG: print(
            f"DEBUG (scrape_nomor): URL tidak valid atau data tidak lengkap, skipping: {url}")
        return None

    if not CLOUDSCAPER_AVAILABLE:
        if is_debug_target and ENABLE_DETAILED_DEBUG: print(
            f"DEBUG (scrape_nomor): cloudscraper tidak tersedia, menggunakan requests biasa (tanpa retry).")
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
            time.sleep(random.uniform(0.7, 1.7))
            response = requests.get(url, headers=headers, timeout=15)
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - fallback requests): Status Code: {response.status_code} untuk {url}")
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")
            postal_code_tag = soup.find("a", class_="ktw")
            if postal_code_tag and postal_code_tag.string:
                code = postal_code_tag.string.strip()
                if code.isdigit() and len(code) == 5: return code
            return None
        except requests.exceptions.RequestException as e:
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - fallback requests): RequestException: {e}")
            return "Error Request (Fallback)"

    scraper = cloudscraper.create_scraper()
    last_error_message = "Gagal setelah beberapa percobaan (nomor.net)"

    for attempt in range(MAX_RETRIES_NOMOR_NET):
        time.sleep(random.uniform(0.5, 1.2))
        if is_debug_target and ENABLE_DETAILED_DEBUG:
            print(f"\n--- DEBUG: scrape_nomor (cloudscraper, Percobaan {attempt + 1}/{MAX_RETRIES_NOMOR_NET}) ---")
            print(f"Mencoba scrape dari URL: {url}")
        try:
            response = scraper.get(url, timeout=25)
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - cloudscraper): Status Code: {response.status_code} untuk {url}")
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")

            code_found_this_attempt = None
            postal_code_tag = soup.find("a", class_="ktw")
            if postal_code_tag and postal_code_tag.string:
                code = postal_code_tag.string.strip()
                if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                    f"DEBUG (scrape_nomor - cloudscraper): Ditemukan tag <a class='ktw'>: '{code}'")
                if code.isdigit() and len(code) == 5:
                    code_found_this_attempt = code
                elif is_debug_target and ENABLE_DETAILED_DEBUG:
                    print(f"DEBUG (scrape_nomor - cloudscraper): Kode dari ktw tidak valid: '{code}'")

            if not code_found_this_attempt:
                if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                    f"DEBUG (scrape_nomor - cloudscraper): Tag <a class='ktw'> tidak ditemukan/valid. Mencoba fallback teks...")
                all_text_nodes = soup.find_all(string=True)
                for text_node in all_text_nodes:
                    stripped_text = str(text_node).strip()
                    if stripped_text.isdigit() and len(stripped_text) == 5:
                        parent_context = text_node.parent.get_text(separator=" ",
                                                                   strip=True).lower() if text_node.parent else ""
                        if "kodepos" in parent_context or len(parent_context) < 100 or "kode pos" in parent_context:
                            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                                f"DEBUG (scrape_nomor - cloudscraper): Fallback - kode pos valid dari teks: {stripped_text}")
                            code_found_this_attempt = stripped_text
                            break

            if code_found_this_attempt:
                return code_found_this_attempt

            last_error_message = "Tidak ditemukan kode pos (setelah parse)"
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - cloudscraper, Percobaan {attempt + 1}): {last_error_message}")

        except requests.exceptions.HTTPError as http_err:
            last_error_message = f"Error HTTP ({http_err.response.status_code})"
            if http_err.response.status_code == 403:
                last_error_message = "Error 403 (Cloudflare block)"
            elif http_err.response.status_code == 404:
                if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                    f"DEBUG (scrape_nomor - cloudscraper): HTTPError 404 untuk {url}. Tidak coba lagi.")
                return "Halaman tidak ditemukan (404)"
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - cloudscraper, Percobaan {attempt + 1}): HTTPError: {http_err}")

        except Exception as e:
            last_error_message = "Error Lainnya (cloudscraper)"
            if is_debug_target and ENABLE_DETAILED_DEBUG: print(
                f"DEBUG (scrape_nomor - cloudscraper, Percobaan {attempt + 1}): Exception: {e}")

        if attempt < MAX_RETRIES_NOMOR_NET - 1:
            if is_debug_target and ENABLE_DETAILED_DEBUG:
                print(
                    f"DEBUG (scrape_nomor - cloudscraper): Percobaan {attempt + 1} gagal ({last_error_message}). Jeda {RETRY_DELAY_NOMOR_NET} detik...")
            time.sleep(RETRY_DELAY_NOMOR_NET)
        else:
            if is_debug_target and ENABLE_DETAILED_DEBUG:
                print(
                    f"DEBUG (scrape_nomor - cloudscraper): Semua ({MAX_RETRIES_NOMOR_NET}) percobaan gagal untuk {url}. Error terakhir: {last_error_message}")
            return last_error_message

    return last_error_message


# Fungsi buat simpen data ke batch file
def save_batch_data(df_slice_to_save, batch_number_val):
    batch_file = os.path.join(batches_dir, f"villages_batch_{batch_number_val}.xlsx")
    try:
        df_slice_to_save.to_excel(batch_file, index=False)
    except Exception as e:
        print(f"--- ERROR (save_batch_data): Gagal saat df_slice_to_save.to_excel() untuk {batch_file}: {e} ---")
        return

    try:
        wb = load_workbook(batch_file)
        ws = wb.active

        for idx_df, row_in_df_for_style in df_slice_to_save.iterrows():
            excel_data_row = df_slice_to_save.index.get_loc(idx_df) + 2

            kode_pos_value_for_styling = str(row_in_df_for_style.get("Kode Pos (Postal Code)", ""))

            is_invalid = False
            if not (kode_pos_value_for_styling.isdigit() and len(kode_pos_value_for_styling) == 5):
                is_invalid = True

            if is_invalid:
                for col_idx_ws in range(1, ws.max_column + 1):
                    ws.cell(row=excel_data_row, column=col_idx_ws).fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        wb.save(batch_file)
        print(
            f"Batch {batch_number_val} (kumulatif untuk run chunk {MY_PROCESS_ID}) disimpan dan di-style di {batch_file}")
    except Exception as e:
        print(f"Error saat styling batch file {batch_file}: {str(e)}")


# --- BAGIAN UTAMA SKRIP ---
URL_KODEPOS_COL = "URL Pencarian kodepos.co.id"
URL_NOMOR_COL = "URL Pencarian nomor.net"
KODE_POS_RESULT_COL = "Kode Pos (Postal Code)"

# Kolom URL dan Kode Pos akan dibuat/di-overwrite di DataFrame `df` (yang sudah di-slice)
print(f"Membuat/Memastikan kolom '{URL_KODEPOS_COL}' di DataFrame yang diproses (ID: {MY_PROCESS_ID})...")
df[URL_KODEPOS_COL] = df.apply(generate_kodepos_url, axis=1)

print(f"Membuat/Memastikan kolom '{URL_NOMOR_COL}' di DataFrame yang diproses (ID: {MY_PROCESS_ID})...")
df[URL_NOMOR_COL] = df.apply(generate_nomor_url, axis=1)

if KODE_POS_RESULT_COL not in df.columns:
    df[KODE_POS_RESULT_COL] = ""
else:
    df[KODE_POS_RESULT_COL] = ""  # Kosongkan untuk diisi ulang oleh proses scraping

print(f"\nContoh URL kodepos.co.id (5 pertama dari data chunk {MY_PROCESS_ID}):")
print(df[URL_KODEPOS_COL].head().tolist())
print(f"\nContoh URL nomor.net (5 pertama dari data chunk {MY_PROCESS_ID}):")
print(df[URL_NOMOR_COL].head().tolist())
print("-" * 30)

batch_size_val = int(BATCH_SIZE) if str(BATCH_SIZE).isdigit() else 20
batch_number = 1

print(
    f"\nMemulai scraping kode pos untuk {len(df)} baris (ID: {MY_PROCESS_ID}, Jatah: {MY_START_ROW_ABSOLUTE} s.d. {MY_END_ROW_EXCLUSIVE_ABSOLUTE - 1} dari file asli)...")
# Loop utama sekarang menggunakan DataFrame `df` yang sudah di-slice
for i in tqdm(range(len(df)), desc=f"Scraping Progress (ID: {MY_PROCESS_ID})"):
    url_kodepos_current = df.at[i, URL_KODEPOS_COL]
    url_nomor_current = df.at[i, URL_NOMOR_COL]

    current_postal_code_found = None

    is_current_row_debug_target = False
    if ENABLE_DETAILED_DEBUG:
        try:
            vil_excel_check = str(df.at[i, "Nama Desa (Village Name)"]).lower().strip() if pd.notna(
                df.at[i, "Nama Desa (Village Name)"]) else ""
            dist_excel_check = str(df.at[i, "Kecamatan (District)"]).lower().strip() if pd.notna(
                df.at[i, "Kecamatan (District)"]) else ""
            is_current_row_debug_target = (
                        DEBUG_TARGET_VILLAGE in vil_excel_check and DEBUG_TARGET_DISTRICT in dist_excel_check)
        except KeyError:
            if i == 0: tqdm.write(
                f"Peringatan: Kolom untuk Debug Target ('Nama Desa (Village Name)' atau 'Kecamatan (District)') tidak ditemukan. Debug per baris tidak aktif.")
            ENABLE_DETAILED_DEBUG = False

    if is_current_row_debug_target and ENABLE_DETAILED_DEBUG:
        # Hitung indeks asli untuk keperluan debug print
        # Ini asumsi df.index adalah 0-based untuk slice. Jika df hasil iloc mempertahankan index asli, maka df.index[i] akan jadi index asli
        # Karena kita .reset_index(drop=True), maka i adalah index di slice. Index asli = MY_START_ROW_ABSOLUTE + i
        original_index_for_debug = MY_START_ROW_ABSOLUTE + i
        print(
            f"\n--- PROCESSING DEBUG TARGET ROW (ID: {MY_PROCESS_ID}, Indeks Asli Excel: {original_index_for_debug + 1}, Indeks Slice: {i}) ---")
        print(f"Desa: {vil_excel_check}, Kecamatan: {dist_excel_check}")
        print(f"URL Nomor.net: {url_nomor_current}")
        print(f"URL Kodepos: {url_kodepos_current}")

    nomor_res = scrape_nomor(url_nomor_current, is_debug_target=is_current_row_debug_target)
    if nomor_res and isinstance(nomor_res, str) and nomor_res.isdigit() and len(nomor_res) == 5:
        current_postal_code_found = nomor_res
    else:
        if is_current_row_debug_target and ENABLE_DETAILED_DEBUG: print(
            f"DEBUG: Hasil nomor.net: {nomor_res}. Mencoba kodepos.co.id...")
        kodepos_res = scrape_kodepos(url_kodepos_current)
        if kodepos_res and isinstance(kodepos_res, str) and kodepos_res.isdigit() and len(kodepos_res) == 5:
            current_postal_code_found = kodepos_res
        elif kodepos_res:
            current_postal_code_found = kodepos_res
        elif nomor_res:
            current_postal_code_found = nomor_res

    df.loc[
        i, KODE_POS_RESULT_COL] = current_postal_code_found if current_postal_code_found else "Invalid Data (tidak ditemukan)"

    if is_current_row_debug_target and ENABLE_DETAILED_DEBUG:
        print(f"Hasil Kode Pos untuk Debug Target (Indeks Slice {i}): {df.loc[i, KODE_POS_RESULT_COL]}")
        print(f"--- END PROCESSING DEBUG TARGET ROW ---\n")

    # Simpan batch secara kumulatif (kumulatif untuk run chunk ini)
    if (i + 1) % batch_size_val == 0 or ((i + 1) == len(df) and len(df) > 0):
        start_index_cumulative_in_slice = 0
        end_index_cumulative_in_slice = i + 1

        df_cumulative_slice_to_save = df.iloc[start_index_cumulative_in_slice:end_index_cumulative_in_slice].copy()

        if not df_cumulative_slice_to_save.empty:
            save_batch_data(df_cumulative_slice_to_save, batch_number)
        else:
            if ENABLE_DETAILED_DEBUG: print(
                f"--- DEBUG: df_cumulative_slice_to_save KOSONG untuk batch {batch_number}. Tidak ada yang disimpan. ---")
        batch_number += 1

# File output akhir hanya akan berisi hasil dari slice yang diproses oleh chunk ini
print(f"\nMenyimpan hasil dari proses ID: {MY_PROCESS_ID} ({len(df)} baris) ke {output_file}...")
try:
    styled_df = df.style.apply(highlight_invalid_rows, axis=1, subset=[KODE_POS_RESULT_COL])
    styled_df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"Selesai! Hasil dari ID: {MY_PROCESS_ID} disimpan di {output_file}")
    print(
        f"PERHATIAN: File ini hanya berisi data untuk jatah ID: {MY_PROCESS_ID} (baris {MY_START_ROW_ABSOLUTE} s.d. {MY_END_ROW_EXCLUSIVE_ABSOLUTE - 1} dari file asli).")
    print(f"Chief dan teman-teman perlu menggabungkannya dengan hasil chunk lain jika sudah semua.")
except Exception as e:
    print(f"Error saat menyimpan file Excel final untuk ID: {MY_PROCESS_ID}: {str(e)}")
    print("Mencoba menyimpan sebagai CSV biasa tanpa styling...")
    try:
        csv_output_file = output_file.replace(".xlsx", "_plain.csv")
        df.to_csv(csv_output_file, index=False)
        print(f"Berhasil menyimpan sebagai CSV biasa ke: {csv_output_file}")
    except Exception as e_csv:
        print(f"Gagal menyimpan sebagai CSV juga: {e_csv}")

